from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
import cloudpickle
import numpy as np
import re
from transformers import BertTokenizer, BertModel
from sklearn.base import BaseEstimator, TransformerMixin
import torch
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook

file_path = r"C:\Users\701540\VS_PY\SENTI\pipeline.pkl"

class TextPreprocessor(BaseEstimator, TransformerMixin):
    def __init__(self, stopwords):
        self.stopwords = stopwords
        self.tokenizer = BertTokenizer.from_pretrained("bert-base-uncased")
        self.bert_model = BertModel.from_pretrained("bert-base-uncased")

    def clean_text(self, text):
        text = re.sub(r"\d+", "", text)
        text = re.sub(r"[^\w\s]", "", text)
        text = text.lower()
        return text
    
    def remove_stopwords(self, text):
        tokens = text.split()
        filtered_tokens = [word for word in tokens if word not in self.stopwords]
        return " ".join(filtered_tokens)
    
    def bert_embedding(self, text):
        inputs = self.tokenizer(text, return_tensors="pt", padding=True, truncation=True)
        with torch.no_grad():
            output= self.bert_model(**inputs)
        embedding = output.last_hidden_state.mean(dim=1).numpy()
        return embedding.flatten()
    
    def transform(self, X, y=None):
        processed_text = []
        for text in X:
            cleaned = self.clean_text(text)
            no_stopword = self.remove_stopwords(cleaned)
            bert_embedded = self.bert_embedding(no_stopword)
            processed_text.append(bert_embedded)
        return np.array(processed_text)
    
    def fit(self, X, y=None):
        return self
    
with open(file_path, "rb") as file:
        senti = cloudpickle.load(file)

#Cors origins
#origins = [
#    "http://172.16.51.223:8001",
#    "http://localhost:8001",
#    "http://localhost",
#    "null"
#]

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins = ["*"],
    allow_credentials = True,
    allow_methods = ["*"],
    allow_headers = ["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")

class SentimentRequest(BaseModel):
    text: str

@app.get("/", response_class=HTMLResponse)
async def read_root():
    with open("static/index.html") as f:
        html_content = f.read()
    return HTMLResponse(content=html_content)

@app.post("/feedback/predict")
async def predict_sentiment(request: SentimentRequest):
    text = request.text
    prediction = senti.predict([text])[0]

    senti_map = {
        "NEGATIVE":"WE APOLOGISE FOR INCONVENIENCE HAPPENED. WE ENSURE FOR BETTER CUSTOMER SERVICE IN FUTURE. THANKS FOR SHOPPING",
        "POSITIVE":"WOW ! WE APPRECIATE YOUR FEEDBACK AND DELIGHTED TO KNOW THAT YOU ENJOYED OUR PRODUCT AND CUSTOMER SERVICE. HAPPY SHOPPING",
        "NEUTRAL": "THANKS FOR THE FEEDBACK. HAPPY SHOPPING"
    }

    sentiment = senti_map.get(prediction, "UNKNOWN")

    data = {
        "Timestamp" : [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        "Text": [text],
        "Prediction": [prediction],
        "Feedback": [sentiment]
    }

    df = pd.DataFrame(data)

    excel_file_path = os.path.abspath("static/prediction_log.xlsx")
    file_exists = os.path.isfile(excel_file_path)

    if file_exists:

        book = load_workbook(excel_file_path)

        with pd.ExcelWriter(excel_file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            writer._book = book

            if "Predictions" in writer._book.sheetnames:
                start_row = writer._book["Predictions"].max_row
                df.to_excel(writer, index=False, header=False, sheet_name="Predictions", startrow=start_row)
            else:
                df.to_excel(writer, index=False, sheet_name="Predictions")
    else:
        with pd.ExcelWriter(excel_file_path, mode="w", engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Predictions")  

    return {"Feedback": sentiment}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)